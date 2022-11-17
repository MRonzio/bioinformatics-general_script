#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 14 9:57:52 2022

merge multiple input text files into a single multisheet excel one,
with each sheet names as file name

@author: MRonzio
"""

import os
import pandas as pd
import argparse
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# COMMAND LINE OPTIONS
def cliopt():
    parser = argparse.ArgumentParser(description='multiple text files merger in excel')
    parser.add_argument('-o', '--outfile', dest='outfilename',
     default='merge.xls', help='output excel filename including extension')
    parser.add_argument('-i', '--inputfiles', dest='files',
     help='input files to merge',nargs='*',required=True)
    parser.add_argument('-r', '--remove-extension', dest='rmext',
     help='remove extension from filename in sheetnames', action='store_true')
    parser.add_argument('-s', '--sep', dest='separator',
     help='if multiple columns: separator character, default is tab, i.e. \t', default="\t")
    parser.add_argument('-ow', '--overwrite', dest='overwrite',action='store_true',
     help='overwrite any existing file with same name')
    return parser.parse_args()

# parse
opt=cliopt()
files = opt.files
separator=opt.separator
overwrite=bool(opt.overwrite)

# remove path and/or filename extension in excel sheets
def clean_name(fullname):
    rmext_bool=bool(opt.rmext)
    if rmext_bool==True:
        filename=os.path.splitext(fullname)[0]
        print(f'filename without extension {filename}')
    else:
        filename=fullname
        print(f'filename with extension {filename}')
    return(os.path.basename(filename))

# Create and save excel file
workbook = openpyxl.Workbook()
for file in files:
    if os.path.isfile(file):
        filename=clean_name(fullname=file)
        workbook.create_sheet(filename)
        print(f"Created {filename} excel sheet")
        df=pd.read_table(file,sep=separator, header=None)
        for r in dataframe_to_rows(df, index=False, header=False):
            workbook[filename].append(r)


workbook.remove(workbook['Sheet'])

if os.path.isfile(opt.outfilename):
    print(f"{opt.outfilename} Already exists")
    if overwrite==True:
        print(f"You set --overwrite so it will be overwritten.")
        workbook.save(opt.outfilename)
        print("done")
    else:
        print(f"Could not save the file, change file name or rerun with --overwrite.")
else:
    workbook.save(opt.outfilename)
    
